"""Smart Excel import for Bad Feedback records.

Two-phase flow keeps the upload disposable:

1. ``inspect_workbook`` — parse the uploaded xlsx in memory (nothing is
   persisted), return sheet names, the active sheet's header row, the
   first data rows as display strings and best-guess suggestions that
   map canonical field keys (``date``, ``source``, ...) to headers.
2. ``commit_import`` — the client re-uploads the SAME file with the
   user-confirmed mapping; rows become ``BadFeedback`` records
   (status=pending), unknown agent names become placeholder users
   (Support or Sales role by the mapping), and the response is a
   per-row report.

Dates arrive as real Excel dates, ISO strings or human text ("12 Mar",
"12.03") with the year often missing — missing years default to the
current one. Agent cells hold comma/semicolon/pipe/newline-separated
lists; each name is resolved by email (case-insensitive) or nickname,
unknown names are slugified (Cyrillic transliterated) into
``nickname@ALLOWED_DOMAIN`` placeholder accounts.
"""

import re
import unicodedata
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    AgentKindEnum,
    BadFeedback,
    BadFeedbackAgent,
    ReviewStatusEnum,
    RoleEnum,
    User,
)
from app.schemas.bad_feedback import (
    ImportCommitResponse,
    ImportInspectResponse,
)
from app.services import user_service

# Canonical importable fields. ``support_agents``/``sales_agents`` are
# the two agent-list columns; everything else is a scalar field.
FIELD_KEYS = (
    "date",
    "source",
    "customer_info",
    "customer_feedback",
    "related_case",
    "support_agents",
    "sales_agents",
)

# Human labels used by the mapping UI, keyed like FIELD_KEYS.
FIELD_LABELS = {
    "date": "Date",
    "source": "Source",
    "customer_info": "Customer's info (CRM link / email)",
    "customer_feedback": "Customer's feedback",
    "related_case": "RA related case",
    "support_agents": "RA Support agents assigned",
    "sales_agents": "Sales agents assigned",
}

# Regexes tried in order by _parse_date_text.
_DATE_PATTERNS = (
    # 2024-03-12 / 2024.03.12 / 2024/03/12
    (re.compile(r"^(\d{4})[-./](\d{1,2})[-./](\d{1,2})$"), "ymd"),
    # 12.03.2024 / 12-03-2024 / 12/03/2024 (day-first per team practice)
    (re.compile(r"^(\d{1,2})[-./](\d{1,2})[-./](\d{4})$"), "dmy"),
    # 12.03 / 12-03 / 12/03 — no year -> current year
    (re.compile(r"^(\d{1,2})[-./](\d{1,2})$"), "dm"),
    # 12 Mar 2024 / 12 March / "12 мар"
    (re.compile(
        r"^(\d{1,2})[\s.-]*([A-Za-zА-Яа-я]{3,9})\.?(?:[\s,.-]+(\d{4}))?$"
    ), "dmon"),
    # Mar 12 / March 12, 2024 / "мар 12"
    (re.compile(
        r"^([A-Za-zА-Яа-я]{3,9})\.?(?:[\s,.-]+(\d{1,2}))?(?:[\s,.-]+(\d{4}))?$"
    ), "mond"),
)

_MONTHS = {
    m: i
    for i, m in enumerate(
        (
            "jan", "feb", "mar", "apr", "may", "jun",
            "jul", "aug", "sep", "oct", "nov", "dec",
        ),
        start=1,
    )
}
for _ru, _idx in (
    ("янв", 1), ("фев", 2), ("мар", 3), ("апр", 4), ("май", 5), ("июн", 6),
    ("июл", 7), ("авг", 8), ("сен", 9), ("окт", 10), ("ноя", 11), ("дек", 12),
):
    _MONTHS[_ru] = _idx

# Agent-cell separators: comma, semicolon, pipe, newline.
_AGENT_SPLIT_RE = re.compile(r"[,;|\n\r]+")

# Cyrillic -> Latin, for slugifying agent names into nicknames.
# Team practice: sheet names are plain English nicknames already, so
# this map only exists as a safety net for stray non-Latin characters.
_CYRILLIC_MAP = str.maketrans({
    "а": "a", "б": "b", "в": "v", "г": "g", "ґ": "g", "д": "d",
    "е": "e", "є": "ie", "ж": "zh", "з": "z", "и": "i", "і": "i",
    "ї": "i", "й": "i", "к": "k", "л": "l", "м": "m", "н": "n",
    "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
    "ф": "f", "х": "kh", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "shch",
    "ы": "y", "э": "e", "ю": "iu", "я": "ia", "ё": "e",
})

# Nickname chars kept verbatim; everything else collapses to '-'.
_SLUG_KEEP_RE = re.compile(r"[^a-z0-9._-]+")

# Safe local parts can be used directly as nicknames when the cell
# holds an email.
_NICKNAME_SAFE_RE = re.compile(r"^[a-z0-9._-]{1,50}$")

_MAX_ROWS = 1000
_PREVIEW_ROWS = 5
_MAX_CELL = 2000

# Row-level skip reasons.
REASON_NO_AGENTS = "no agents in the row"
REASON_BAD_DATE = "unparsable date (fix it in the sheet or clear the cell)"


def _display(value: Any) -> str | None:
    """Normalize any Excel cell to a trimmed display string or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _parse_date_text(text: str, today: date) -> date | None:
    """Parse a human date string; missing year -> ``today.year``.

    Day-first for numeric forms (team practice: "12.03" is March 12).
    Returns None when nothing matches — the row is skipped with a
    reason rather than silently mis-dated.
    """
    text = text.strip().strip('"').strip()
    if not text:
        return None
    lowered = text.lower().replace(",", " ")
    for pattern, shape in _DATE_PATTERNS:
        match = pattern.match(lowered)
        if not match:
            continue
        groups = match.groups()
        try:
            if shape == "ymd":
                y, m, d = int(groups[0]), int(groups[1]), int(groups[2])
            elif shape == "dmy":
                d, m, y = int(groups[0]), int(groups[1]), int(groups[2])
            elif shape == "dm":
                d, m, y = int(groups[0]), int(groups[1]), today.year
            elif shape == "dmon":
                d = int(groups[0])
                m = _MONTHS.get(groups[1][:3])
                if m is None:
                    continue
                y = int(groups[2]) if groups[2] else today.year
            else:  # mond
                m = _MONTHS.get(groups[0][:3])
                if m is None:
                    continue
                d = int(groups[1]) if groups[1] else 1
                y = int(groups[2]) if groups[2] else today.year
        except (TypeError, ValueError):
            continue
        try:
            return date(y, m, d)
        except ValueError:
            continue
    return None


def parse_feedback_date(value: Any, today: date | None = None) -> date | None:
    """Coerce an Excel cell into a date.

    Real dates/datetimes pass through; strings go through
    :func:`_parse_date_text` (missing year -> current year). Unparsable
    text returns None.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return _parse_date_text(str(value), today or date.today())


def _split_agents(raw: str) -> list[str]:
    """Split an agent cell into cleaned name tokens."""
    tokens = (t.strip(" .\u00a0\t") for t in _AGENT_SPLIT_RE.split(raw))
    return [t for t in tokens if t]


def _slugify_name(raw: str) -> str:
    """Turn an agent display name into a valid placeholder nickname.

    Transliterates Cyrillic, keeps ``[a-z0-9._-]`` (anything else ->
    '-'), collapses separators, trims to 50 chars. Returns '' when
    nothing usable survives.
    """
    text = unicodedata.normalize("NFKD", raw.strip().lower())
    translit = text.translate(_CYRILLIC_MAP)
    slug = _SLUG_KEEP_RE.sub("-", translit).strip("-.")
    slug = re.sub(r"-{2,}", "-", slug)
    return slug[:50]


def kind_to_role(kind: AgentKindEnum) -> RoleEnum:
    """Agent kind -> the role a synthesized account receives."""
    return (
        RoleEnum.SUPPORT if kind is AgentKindEnum.SUPPORT else RoleEnum.SALES
    )


def _ensure_frontline_role(user: User, kind: AgentKindEnum) -> None:
    """Grant a found user the front-line role implied by the column.

    Front-line roles are additive badges: a QA-only account referenced
    by a sheet gains Support/Sales without losing anything, and a user
    already holding the OTHER front-line role gains this one too.
    """
    role = kind_to_role(kind)
    if role not in user.roles:
        user.roles = [*user.roles, role]


def _resolve_or_create_user(
    raw_name: str,
    kind: AgentKindEnum,
    known_by_email: dict[str, User],
    known_by_nick: dict[str, User],
    created_log: list[str],
) -> User | None:
    """Find a user by email/nickname or synthesize a placeholder.

    Lookup: full email (case-insensitive) -> nickname slug. Unknown
    names become ``{slug}@ALLOWED_DOMAIN`` accounts with the role
    implied by the column (Support/Sales).
    """
    lowered = raw_name.strip().lower()
    nick = _slugify_name(raw_name)
    user = known_by_email.get(lowered)
    if user is None and "@" in lowered:
        local = lowered.split("@", 1)[0]
        local_nick = local if _NICKNAME_SAFE_RE.fullmatch(local) else _slugify_name(local)
        user = known_by_nick.get(local_nick)
    if user is None and nick:
        user = known_by_nick.get(nick)
    if user is not None:
        _ensure_frontline_role(user, kind)
        return user

    if not nick:
        return None
    email = user_service.placeholder_email(nick)
    user = User(email=email, name=None, roles=[kind_to_role(kind)])
    created_log.append(f"{raw_name.strip()} -> {email}")
    known_by_email[email] = user
    known_by_nick.setdefault(nick, user)
    return user


async def _load_user_maps(
    db: AsyncSession,
) -> tuple[dict[str, User], dict[str, User]]:
    """Load every non-deleted user into email/nickname lookup maps."""
    rows = (await db.execute(select(User).where(User.active_filter()))).scalars()
    by_email: dict[str, User] = {}
    by_nick: dict[str, User] = {}
    for user in rows:
        by_email[user.email.lower()] = user
        by_nick[user.email.split("@", 1)[0].lower()] = user
    return by_email, by_nick


def _header_index(headers: list[str], wanted: str | None) -> int | None:
    """Map a chosen header text to its column index (None = unmapped)."""
    if not wanted:
        return None
    wanted = wanted.strip()
    for idx, header in enumerate(headers):
        if header.strip() == wanted:
            return idx
    return None


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    """Best-guess canonical field -> header text from the header row."""
    lowered = [h.lower() for h in headers]
    suggestions: dict[str, str] = {}
    rules = (
        ("date", ("date", "дата")),
        ("source", ("source", "источник", "channel")),
        ("customer_info", ("customer", "клиент", "email", "crm")),
        ("customer_feedback", ("feedback", "comment", "отзыв")),
        ("related_case", ("case", "кейс", "ticket")),
        ("support_agents", ("support agent", "support", "ra support")),
        ("sales_agents", ("sales",)),
    )
    used: set[int] = set()
    for key, needles in rules:
        found = None
        for idx, header in enumerate(lowered):
            if idx in used:
                continue
            if any(header.startswith(n) for n in needles):
                found = idx
                break
        if found is None:
            for idx, header in enumerate(lowered):
                if idx in used:
                    continue
                if any(n in header for n in needles):
                    found = idx
                    break
        if found is not None:
            used.add(found)
            suggestions[key] = headers[found]
    return suggestions


def inspect_workbook(content: bytes) -> ImportInspectResponse:
    """Parse the uploaded workbook in memory (nothing persisted)."""
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            raise ValueError("The first sheet is empty.")
        headers = [
            _display(value) or f"Column {idx + 1}"
            for idx, value in enumerate(header_row)
        ]
        data_rows: list[list[str]] = []
        total = 0
        for values in rows:
            if not any(v is not None and str(v).strip() for v in values):
                continue
            total += 1
            if len(data_rows) < _PREVIEW_ROWS:
                data_rows.append(
                    [_display(v) or "" for v in values[: len(headers)]]
                )
        return ImportInspectResponse(
            sheet_names=list(workbook.sheetnames),
            active_sheet=sheet.title,
            headers=headers,
            suggestions=suggest_mapping(headers),
            preview=data_rows,
            total_rows=total,
        )
    finally:
        workbook.close()


async def commit_import(
    db: AsyncSession,
    content: bytes,
    mapping: dict[str, str | None],
    assigned_qa_id: int | None,
    created_by: int,
) -> ImportCommitResponse:
    """Create pending BadFeedback rows from the mapped sheet.

    ``mapping`` keys are :data:`FIELD_KEYS` with header texts (or
    None/""). Rows lacking ANY agent name are skipped with a reason;
    unknown agent names are created as placeholder users; everything
    commits atomically.
    """
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            raise ValueError("The first sheet is empty.")
        headers = [
            _display(value) or f"Column {idx + 1}"
            for idx, value in enumerate(header_row)
        ]
        columns = {
            key: _header_index(headers, mapping.get(key))
            for key in FIELD_KEYS
        }

        by_email, by_nick = await _load_user_maps(db)
        today = date.today()
        created: list[BadFeedback] = []
        skipped: list[tuple[int, str]] = []
        created_users: list[str] = []

        for row_number, values in enumerate(rows, start=2):
            if row_number - 2 >= _MAX_ROWS:
                skipped.append((row_number, "row limit reached"))
                continue
            if not any(v is not None and str(v).strip() for v in values):
                continue

            def cell(key: str) -> Any:
                idx = columns.get(key)
                if idx is None or idx >= len(values):
                    return None
                return values[idx]

            def text(key: str) -> str | None:
                value = _display(cell(key))
                if value is not None and len(value) > _MAX_CELL:
                    value = value[:_MAX_CELL]
                return value

            raw_date = cell("date")
            fb_date = parse_feedback_date(raw_date, today)
            if raw_date not in (None, "") and fb_date is None:
                skipped.append((row_number, REASON_BAD_DATE))
                continue

            agent_rows: list[tuple[User, AgentKindEnum]] = []
            for key, kind in (
                ("support_agents", AgentKindEnum.SUPPORT),
                ("sales_agents", AgentKindEnum.SALES),
            ):
                raw = text(key)
                if not raw:
                    continue
                for name in _split_agents(raw):
                    user = _resolve_or_create_user(
                        name, kind, by_email, by_nick, created_users
                    )
                    if user is not None:
                        agent_rows.append((user, kind))
            if not agent_rows:
                skipped.append((row_number, REASON_NO_AGENTS))
                continue

            feedback = BadFeedback(
                fb_date=fb_date,
                source=text("source"),
                customer_info=text("customer_info"),
                customer_feedback=text("customer_feedback"),
                related_case=text("related_case"),
                status=ReviewStatusEnum.PENDING,
                assigned_qa_id=assigned_qa_id,
                created_by=created_by,
            )
            # Dedup by email, not by user.id: rows created in this
            # import have no id until the flush, so two brand-new
            # agents would collapse into one "None" key.
            seen: set[tuple[str, AgentKindEnum]] = set()
            for user, kind in agent_rows:
                pair = (user.email.lower(), kind)
                if pair in seen:
                    continue
                seen.add(pair)
                feedback.agents.append(BadFeedbackAgent(user=user, kind=kind))
            created.append(feedback)
            db.add(feedback)

        await db.commit()
        return ImportCommitResponse(
            created=[
                {
                    "id": fb.id,
                    "related_case": fb.related_case,
                    "agent_labels": [a.user.nickname for a in fb.agents],
                }
                for fb in created
            ],
            skipped=[
                {"row_number": n, "reason": reason}
                for n, reason in skipped
            ],
            created_users=created_users,
            created_count=len(created),
            skipped_count=len(skipped),
        )
    finally:
        workbook.close()
